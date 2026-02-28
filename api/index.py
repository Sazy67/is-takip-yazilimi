from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import os
from datetime import datetime, timedelta
import secrets
import jwt
import psycopg2
from psycopg2.extras import RealDictCursor
from openpyxl import Workbook, load_workbook
from io import BytesIO

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(16))
CORS(app, supports_credentials=True, origins=['*'])

# Neon Postgres bağlantısı
DATABASE_URL = os.environ.get('POSTGRES_URL')

def get_db():
    conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
    return conn

def init_db():
    conn = get_db()
    cur = conn.cursor()
    
    # Users tablosu
    cur.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            username VARCHAR(50) UNIQUE NOT NULL,
            password VARCHAR(100) NOT NULL,
            role VARCHAR(20) NOT NULL
        )
    ''')
    
    # Kayitlar tablosu
    cur.execute('''
        CREATE TABLE IF NOT EXISTS kayitlar (
            id SERIAL PRIMARY KEY,
            bolum VARCHAR(100),
            teklif_no VARCHAR(50),
            musteri_ismi VARCHAR(200),
            teklif_tarihi DATE,
            onay_tarihi DATE,
            uretime_verilme_tarihi DATE,
            uretim_numarasi VARCHAR(50),
            cam_siparis_tarihi DATE,
            cam_siparis_numarasi VARCHAR(50),
            cam_adedi VARCHAR(50),
            uretim_planlama_tarihi DATE,
            paketleme_tarihi DATE,
            kasetleme_tarihi DATE,
            sevk_tarihi DATE,
            teklif_durumu BOOLEAN DEFAULT FALSE,
            imalat_durumu BOOLEAN DEFAULT FALSE,
            notlar TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Mevcut tabloya yeni kolonları ekle (eğer yoksa)
    try:
        cur.execute('''
            ALTER TABLE kayitlar 
            ADD COLUMN IF NOT EXISTS teklif_durumu BOOLEAN DEFAULT FALSE,
            ADD COLUMN IF NOT EXISTS imalat_durumu BOOLEAN DEFAULT FALSE
        ''')
        conn.commit()
    except:
        pass
    
    # Activity logs tablosu
    cur.execute('''
        CREATE TABLE IF NOT EXISTS activity_logs (
            id SERIAL PRIMARY KEY,
            username VARCHAR(50),
            action VARCHAR(100),
            details TEXT,
            ip_address VARCHAR(50),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    conn.commit()
    
    # Varsayılan kullanıcıları kontrol et ve sadece yoksa ekle
    cur.execute("SELECT COUNT(*) as count FROM users")
    user_count = cur.fetchone()['count']
    
    if user_count == 0:
        try:
            cur.execute("INSERT INTO users (username, password, role) VALUES (%s, %s, %s)", ('admin', 'admin123', 'admin'))
            cur.execute("INSERT INTO users (username, password, role) VALUES (%s, %s, %s)", ('user', 'user123', 'user'))
            conn.commit()
        except:
            pass
    
    # Örnek kayıtları kontrol et ve sadece yoksa ekle
    cur.execute("SELECT COUNT(*) as count FROM kayitlar")
    kayit_count = cur.fetchone()['count']
    
    if kayit_count == 0:
        try:
            cur.execute('''
                INSERT INTO kayitlar (
                    bolum, teklif_no, musteri_ismi, teklif_tarihi, onay_tarihi,
                    uretime_verilme_tarihi, uretim_numarasi, cam_siparis_tarihi,
                    cam_siparis_numarasi, cam_adedi, uretim_planlama_tarihi,
                    paketleme_tarihi, kasetleme_tarihi, sevk_tarihi, notlar
                ) VALUES 
                (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s),
                (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ''', (
                'Üretim', 'TK-2025-001', 'ABC İnşaat Ltd.', '2025-01-15', '2025-01-20', '2025-01-25', 'UR-001', '2025-01-22', 'CS-001', '50', '2025-01-28', '2025-02-05', '2025-02-08', '2025-02-10', 'İlk sipariş, öncelikli',
                'Satış', 'TK-2025-002', 'XYZ Yapı A.Ş.', '2025-01-18', '2025-01-22', '2025-01-27', 'UR-002', '2025-01-25', 'CS-002', '75', '2025-02-01', '2025-02-08', '2025-02-12', '2025-02-15', 'Standart teslimat'
            ))
            conn.commit()
        except:
            pass
    
    cur.close()
    conn.close()

# İlk çalıştırmada tabloları oluştur
try:
    init_db()
except:
    pass

def create_token(user_id, username, role):
    payload = {
        'user_id': user_id,
        'username': username,
        'role': role,
        'exp': datetime.utcnow() + timedelta(days=7)
    }
    return jwt.encode(payload, app.secret_key, algorithm='HS256')

def verify_token(token):
    try:
        payload = jwt.decode(token, app.secret_key, algorithms=['HS256'])
        return payload
    except:
        return None

def log_activity(username, action, details='', ip_address=''):
    try:
        conn = get_db()
        cur = conn.cursor()
        cur.execute('''
            INSERT INTO activity_logs (username, action, details, ip_address)
            VALUES (%s, %s, %s, %s)
        ''', (username, action, details, ip_address))
        conn.commit()
        cur.close()
        conn.close()
    except:
        pass

@app.route('/api/login', methods=['POST'])
def login():
    data = request.json
    username = data.get('username')
    password = data.get('password')
    ip_address = request.remote_addr
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT * FROM users WHERE username=%s AND password=%s', (username, password))
    user = cur.fetchone()
    cur.close()
    conn.close()
    
    if user:
        token = create_token(user['id'], user['username'], user['role'])
        log_activity(username, 'Giriş Yaptı', f'{user["role"]} olarak giriş', ip_address)
        return jsonify({
            'success': True,
            'token': token,
            'user': {
                'username': user['username'],
                'role': user['role']
            }
        })
    
    log_activity(username or 'Bilinmeyen', 'Başarısız Giriş', 'Hatalı kullanıcı adı veya şifre', ip_address)
    return jsonify({'success': False, 'message': 'Kullanıcı adı veya şifre hatalı'}), 401

@app.route('/api/logout', methods=['POST'])
def logout():
    return jsonify({'success': True})

@app.route('/api/me', methods=['GET'])
def get_current_user():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user_data = verify_token(token)
    
    if user_data:
        return jsonify({
            'username': user_data['username'],
            'role': user_data['role']
        })
    return jsonify({'error': 'Not authenticated'}), 401

@app.route('/api/kayitlar', methods=['GET'])
def get_kayitlar():
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT * FROM kayitlar ORDER BY id DESC')
    kayitlar = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify(kayitlar)

@app.route('/api/kayitlar', methods=['POST'])
def create_kayit():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user_data = verify_token(token)
    
    if not user_data or user_data['role'] != 'admin':
        return jsonify({'error': 'Yetkisiz işlem'}), 403
    
    data = request.json
    conn = get_db()
    cur = conn.cursor()
    
    # Boş string'leri None'a çevir
    def clean_value(val):
        return None if val == '' else val
    
    cur.execute('''
        INSERT INTO kayitlar (
            bolum, teklif_no, musteri_ismi, teklif_tarihi, onay_tarihi,
            uretime_verilme_tarihi, uretim_numarasi, cam_siparis_tarihi,
            cam_siparis_numarasi, cam_adedi, uretim_planlama_tarihi,
            paketleme_tarihi, kasetleme_tarihi, sevk_tarihi, notlar
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id
    ''', (
        clean_value(data.get('bolum')), clean_value(data.get('teklif_no')), clean_value(data.get('musteri_ismi')),
        clean_value(data.get('teklif_tarihi')), clean_value(data.get('onay_tarihi')), clean_value(data.get('uretime_verilme_tarihi')),
        clean_value(data.get('uretim_numarasi')), clean_value(data.get('cam_siparis_tarihi')), clean_value(data.get('cam_siparis_numarasi')),
        clean_value(data.get('cam_adedi')), clean_value(data.get('uretim_planlama_tarihi')), clean_value(data.get('paketleme_tarihi')),
        clean_value(data.get('kasetleme_tarihi')), clean_value(data.get('sevk_tarihi')), clean_value(data.get('notlar'))
    ))
    
    kayit_id = cur.fetchone()['id']
    conn.commit()
    cur.close()
    conn.close()
    
    log_activity(user_data['username'], 'Kayıt Ekledi', f'Yeni kayıt: {data.get("musteri_ismi")} - {data.get("teklif_no")}', request.remote_addr)
    
    return jsonify({'id': kayit_id}), 201

@app.route('/api/kayitlar/<int:id>', methods=['PUT'])
def update_kayit(id):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user_data = verify_token(token)
    
    if not user_data:
        return jsonify({'error': 'Yetkisiz işlem'}), 403
    
    data = request.json
    conn = get_db()
    cur = conn.cursor()
    
    # Kullanıcı sadece not güncelleyebilir
    if user_data['role'] == 'user':
        cur.execute('SELECT notlar FROM kayitlar WHERE id=%s', (id,))
        kayit = cur.fetchone()
        
        if kayit:
            eski_not = kayit['notlar'] or ''
            yeni_not = data.get('notlar', '')
            
            tarih = datetime.now().strftime('%Y-%m-%d %H:%M')
            not_ekleme = f"\n[{user_data['username']} - {tarih}]: {yeni_not}"
            
            guncel_not = eski_not + not_ekleme
            
            cur.execute('UPDATE kayitlar SET notlar=%s WHERE id=%s', (guncel_not, id))
            conn.commit()
            cur.close()
            conn.close()
            log_activity(user_data['username'], 'Not Ekledi', f'Kayıt ID: {id}', request.remote_addr)
            return jsonify({'success': True})
    
    # Admin tüm alanları güncelleyebilir
    if user_data['role'] == 'admin':
        # Boş string'leri None'a çevir
        def clean_value(val):
            return None if val == '' else val
        
        cur.execute('''
            UPDATE kayitlar SET
                bolum=%s, teklif_no=%s, musteri_ismi=%s, teklif_tarihi=%s, onay_tarihi=%s,
                uretime_verilme_tarihi=%s, uretim_numarasi=%s, cam_siparis_tarihi=%s,
                cam_siparis_numarasi=%s, cam_adedi=%s, uretim_planlama_tarihi=%s,
                paketleme_tarihi=%s, kasetleme_tarihi=%s, sevk_tarihi=%s, 
                teklif_durumu=%s, imalat_durumu=%s, notlar=%s
            WHERE id=%s
        ''', (
            clean_value(data.get('bolum')), clean_value(data.get('teklif_no')), clean_value(data.get('musteri_ismi')),
            clean_value(data.get('teklif_tarihi')), clean_value(data.get('onay_tarihi')), clean_value(data.get('uretime_verilme_tarihi')),
            clean_value(data.get('uretim_numarasi')), clean_value(data.get('cam_siparis_tarihi')), clean_value(data.get('cam_siparis_numarasi')),
            clean_value(data.get('cam_adedi')), clean_value(data.get('uretim_planlama_tarihi')), clean_value(data.get('paketleme_tarihi')),
            clean_value(data.get('kasetleme_tarihi')), clean_value(data.get('sevk_tarihi')), 
            data.get('teklif_durumu', False), data.get('imalat_durumu', False), clean_value(data.get('notlar')), id
        ))
        conn.commit()
        cur.close()
        conn.close()
        log_activity(user_data['username'], 'Kayıt Güncelledi', f'Kayıt ID: {id} - {data.get("musteri_ismi")}', request.remote_addr)
        return jsonify({'success': True})
    
    cur.close()
    conn.close()
    return jsonify({'error': 'Yetkisiz işlem'}), 403

@app.route('/api/kayitlar/<int:id>', methods=['DELETE'])
def delete_kayit(id):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user_data = verify_token(token)
    
    if not user_data or user_data['role'] != 'admin':
        return jsonify({'error': 'Yetkisiz işlem'}), 403
    
    conn = get_db()
    cur = conn.cursor()
    
    # Silmeden önce kayıt bilgisini al
    cur.execute('SELECT musteri_ismi, teklif_no FROM kayitlar WHERE id=%s', (id,))
    kayit = cur.fetchone()
    
    cur.execute('DELETE FROM kayitlar WHERE id=%s', (id,))
    conn.commit()
    cur.close()
    conn.close()
    
    if kayit:
        log_activity(user_data['username'], 'Kayıt Sildi', f'Kayıt ID: {id} - {kayit["musteri_ismi"]} ({kayit["teklif_no"]})', request.remote_addr)
    
    return jsonify({'success': True})

# Kullanıcı yönetimi endpoint'leri
@app.route('/api/users', methods=['GET'])
def get_users():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user_data = verify_token(token)
    
    if not user_data or user_data['role'] != 'admin':
        return jsonify({'error': 'Yetkisiz işlem'}), 403
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT id, username, role FROM users ORDER BY id')
    users = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify(users)

@app.route('/api/users', methods=['POST'])
def create_user():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user_data = verify_token(token)
    
    if not user_data or user_data['role'] != 'admin':
        return jsonify({'error': 'Yetkisiz işlem'}), 403
    
    data = request.json
    conn = get_db()
    cur = conn.cursor()
    
    try:
        cur.execute('''
            INSERT INTO users (username, password, role)
            VALUES (%s, %s, %s)
            RETURNING id
        ''', (data.get('username'), data.get('password'), data.get('role')))
        
        user_id = cur.fetchone()['id']
        conn.commit()
        cur.close()
        conn.close()
        
        log_activity(user_data['username'], 'Kullanıcı Oluşturdu', f'Yeni kullanıcı: {data.get("username")} ({data.get("role")})', request.remote_addr)
        
        return jsonify({'id': user_id}), 201
    except Exception as e:
        conn.close()
        return jsonify({'error': 'Kullanıcı adı zaten mevcut'}), 400

@app.route('/api/users/<int:id>', methods=['PUT'])
def update_user(id):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user_data = verify_token(token)
    
    if not user_data or user_data['role'] != 'admin':
        return jsonify({'error': 'Yetkisiz işlem'}), 403
    
    data = request.json
    conn = get_db()
    cur = conn.cursor()
    
    try:
        cur.execute('''
            UPDATE users SET username=%s, password=%s, role=%s
            WHERE id=%s
        ''', (data.get('username'), data.get('password'), data.get('role'), id))
        
        conn.commit()
        cur.close()
        conn.close()
        
        log_activity(user_data['username'], 'Kullanıcı Güncelledi', f'Kullanıcı ID: {id} - {data.get("username")} ({data.get("role")})', request.remote_addr)
        
        return jsonify({'success': True})
    except Exception as e:
        conn.close()
        return jsonify({'error': 'Kullanıcı adı zaten mevcut'}), 400

@app.route('/api/users/<int:id>', methods=['DELETE'])
def delete_user(id):
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user_data = verify_token(token)
    
    if not user_data or user_data['role'] != 'admin':
        return jsonify({'error': 'Yetkisiz işlem'}), 403
    
    # Kendi hesabını silemesin
    if user_data['user_id'] == id:
        return jsonify({'error': 'Kendi hesabınızı silemezsiniz'}), 400
    
    conn = get_db()
    cur = conn.cursor()
    
    # Silmeden önce kullanıcı bilgisini al
    cur.execute('SELECT username, role FROM users WHERE id=%s', (id,))
    user = cur.fetchone()
    
    cur.execute('DELETE FROM users WHERE id=%s', (id,))
    conn.commit()
    cur.close()
    conn.close()
    
    if user:
        log_activity(user_data['username'], 'Kullanıcı Sildi', f'Kullanıcı ID: {id} - {user["username"]} ({user["role"]})', request.remote_addr)
    
    return jsonify({'success': True})

# Gizli Log Sayfası - Sadece özel URL ile erişilebilir
@app.route('/api/secret-logs-x9k2p7m4', methods=['GET'])
def get_secret_logs():
    # Basit güvenlik: Query parameter kontrolü
    secret_key = request.args.get('key')
    if secret_key != 'suat2025':
        return jsonify({'error': 'Unauthorized'}), 403
    
    conn = get_db()
    cur = conn.cursor()
    # Sadece kayıt güncelleme ve not ekleme loglarını göster
    cur.execute('''
        SELECT id, username, action, details, ip_address, created_at 
        FROM activity_logs 
        WHERE action IN ('Kayıt Güncelledi', 'Not Ekledi')
        ORDER BY created_at DESC 
        LIMIT 500
    ''')
    logs = cur.fetchall()
    cur.close()
    conn.close()
    return jsonify(logs)

# Excel Export
@app.route('/api/export-excel', methods=['GET'])
def export_excel():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user_data = verify_token(token)
    
    if not user_data or user_data['role'] != 'admin':
        return jsonify({'error': 'Yetkisiz işlem'}), 403
    
    conn = get_db()
    cur = conn.cursor()
    cur.execute('SELECT * FROM kayitlar ORDER BY id')
    kayitlar = cur.fetchall()
    cur.close()
    conn.close()
    
    # Excel oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Kayıtlar"
    
    # Header
    headers = ['ID', 'Bölüm', 'Teklif No', 'Müşteri İsmi', 'Teklif Tarihi', 'Onay Tarihi',
               'Üretime Verilme Tarihi', 'Üretim Numarası', 'Cam Sipariş Tarihi', 
               'Cam Sipariş Numarası', 'Cam Adedi', 'Üretim Planlama Tarihi',
               'Paketleme Tarihi', 'Kasetleme Tarihi', 'Sevk Tarihi', 
               'Teklif Durumu', 'İmalat Durumu', 'Notlar']
    ws.append(headers)
    
    # Data
    for kayit in kayitlar:
        row = [
            kayit['id'],
            kayit['bolum'],
            kayit['teklif_no'],
            kayit['musteri_ismi'],
            str(kayit['teklif_tarihi']) if kayit['teklif_tarihi'] else '',
            str(kayit['onay_tarihi']) if kayit['onay_tarihi'] else '',
            str(kayit['uretime_verilme_tarihi']) if kayit['uretime_verilme_tarihi'] else '',
            kayit['uretim_numarasi'],
            str(kayit['cam_siparis_tarihi']) if kayit['cam_siparis_tarihi'] else '',
            kayit['cam_siparis_numarasi'],
            kayit['cam_adedi'],
            str(kayit['uretim_planlama_tarihi']) if kayit['uretim_planlama_tarihi'] else '',
            str(kayit['paketleme_tarihi']) if kayit['paketleme_tarihi'] else '',
            str(kayit['kasetleme_tarihi']) if kayit['kasetleme_tarihi'] else '',
            str(kayit['sevk_tarihi']) if kayit['sevk_tarihi'] else '',
            'Evet' if kayit['teklif_durumu'] else 'Hayır',
            'Evet' if kayit['imalat_durumu'] else 'Hayır',
            kayit['notlar']
        ]
        ws.append(row)
    
    # BytesIO'ya kaydet
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    log_activity(user_data['username'], 'Excel Yedek Aldı', f'{len(kayitlar)} kayıt', request.remote_addr)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'kayitlar_yedek_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

# Excel Import
@app.route('/api/import-excel', methods=['POST'])
def import_excel():
    token = request.headers.get('Authorization', '').replace('Bearer ', '')
    user_data = verify_token(token)
    
    if not user_data or user_data['role'] != 'admin':
        return jsonify({'error': 'Yetkisiz işlem'}), 403
    
    if 'file' not in request.files:
        return jsonify({'error': 'Dosya bulunamadı'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Dosya seçilmedi'}), 400
    
    try:
        wb = load_workbook(file)
        ws = wb.active
        
        conn = get_db()
        cur = conn.cursor()
        
        added = 0
        skipped = 0
        total = 0
        
        # İlk satır header, 2. satırdan başla
        for row in ws.iter_rows(min_row=2, values_only=True):
            total += 1
            
            # Boş satırları atla
            if not any(row):
                skipped += 1
                continue
            
            try:
                # Tarih değerlerini düzelt
                def parse_date(val):
                    if not val or val == '':
                        return None
                    if isinstance(val, datetime):
                        return val.date()
                    return str(val)
                
                # Boolean değerleri düzelt
                def parse_bool(val):
                    if isinstance(val, bool):
                        return val
                    if isinstance(val, str):
                        return val.lower() in ['evet', 'yes', 'true', '1']
                    return False
                
                cur.execute('''
                    INSERT INTO kayitlar (
                        bolum, teklif_no, musteri_ismi, teklif_tarihi, onay_tarihi,
                        uretime_verilme_tarihi, uretim_numarasi, cam_siparis_tarihi,
                        cam_siparis_numarasi, cam_adedi, uretim_planlama_tarihi,
                        paketleme_tarihi, kasetleme_tarihi, sevk_tarihi,
                        teklif_durumu, imalat_durumu, notlar
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ''', (
                    row[1], row[2], row[3],  # bolum, teklif_no, musteri_ismi
                    parse_date(row[4]), parse_date(row[5]), parse_date(row[6]),  # tarihler
                    row[7], parse_date(row[8]), row[9],  # uretim_numarasi, cam_siparis_tarihi, cam_siparis_numarasi
                    row[10], parse_date(row[11]), parse_date(row[12]),  # cam_adedi, uretim_planlama_tarihi, paketleme_tarihi
                    parse_date(row[13]), parse_date(row[14]),  # kasetleme_tarihi, sevk_tarihi
                    parse_bool(row[15]), parse_bool(row[16]),  # teklif_durumu, imalat_durumu
                    row[17] if len(row) > 17 else None  # notlar
                ))
                added += 1
            except Exception as e:
                print(f"Row error: {e}")
                skipped += 1
                continue
        
        conn.commit()
        cur.close()
        conn.close()
        
        log_activity(user_data['username'], 'Excel Yedek Yükledi', f'{added} kayıt eklendi, {skipped} atlandı', request.remote_addr)
        
        return jsonify({
            'success': True,
            'total': total,
            'added': added,
            'skipped': skipped
        })
        
    except Exception as e:
        return jsonify({'error': f'Excel işleme hatası: {str(e)}'}), 400

